# 웹 크롤링을 통한 시각화
# conda install beautifulsoup4, selenium

from bs4 import BeautifulSoup
from selenium import webdriver  #chrome webdriver 다운로드 필요
from openpyxl import Workbook
import time

# 웹 크롤링 수행 후 엑셀에 저장
"""
driver = webdriver.Chrome("./chromedriver")
driver.implicitly_wait(3)
result_list = []

for i in range(1,11):
    print( i, "페이지")    
    driver.get( "https://www1.president.go.kr/petitions/best?page=" + str(i) )
    html = driver.page_source
    soup = BeautifulSoup( html, 'html.parser' )

    for j in soup.select("#cont_view > div.cs_area > div > div > div.board.text > div.b_list.category > div.bl_body > ul > li"):
        #print( i.text )
        print( j.find("div", class_ = "bl_subject" ).text[3:].strip() )
        result_list.append( j.find("div", class_ = "bl_subject" ).text[3:].strip() )
    time.sleep(3)
    
driver.close()

write_workbook = Workbook()
write_cell = write_workbook.active
for i in range(1, len(result_list) + 1 ):
    write_cell.cell( i, 1, result_list[i-1] )

write_workbook.save("blue.xlsx")    
write_workbook.close()



# 엑셀에 저장된 결과값 조회

from openpyxl import load_workbook

read_workbook = load_workbook("./blue.xlsx")
read_cell = read_workbook.active
write_list = []

for i in range(1, 150) :
    write_list.append( read_cell.cell(i,1).value )
#print( write_list )    

for i in write_list:
    print(i)
"""
"""
# 자연어 분석
#JDK설치 후 Anaconda Prompt 오픈
#conda install -c conda-forge jpype1
#pip install konlpy

from konlpy.tag import Kkma

kkma = Kkma()
print( kkma.sentences("네 안녕하세요. 반갑습니다. 저는 회사에 다닙니다") )
print( kkma.nouns("우리는 민족중흥의 역사적 사명을 띠고 이땅에 태어났을까? 아니면 커피 혹은 우유") )
"""

# 단어 개수 세기
from konlpy.tag import Kkma
import collections
kkma = Kkma()
"""
list_a = ["가","라","다","가","나","라","마"]
list_b = ["가","다","나","나","가","다","다","다"]
list_c = list_a + list_b
for k, v in collections.Counter(list_c).items() :
    print(k,'는', v , '번 존재합니다')

print( collections.Counter(list_c).most_common(3) )

sentence_list = [
        "우리  축구 야구 축구 축구",
        "우리는 민족중흥의 성악설 역사적 사명을 띠고 축구 이땅에 태어났다",
        "우리 나라 말씀이 중국과 성악 달라 고생이 많다 축구나 할까.",
        "우리 오늘 저녁은 간단하게 깁밥, 라면, 떡볶이, 축구 순대, 어묵, 돈까스로",
        "공자 김밥 맹자 순자 성선설 성악설 종교 예수 성악설 석가",
        "우리 라면, 성악설 떡볶이, 순대, 어묵 예수 석가",
        "우리는 김밥 중국 미국 민족의 역사를 사명, 역사와 축구 사명, 역사, 성악설"
        ]
result_list = []
for i in sentence_list :
    #print( kkma.nouns( i ) )
    result_list += kkma.nouns( i )
    #print ( )
    #collections.Counter(list_c).most_common(3)

print( collections.Counter(result_list).most_common(5) )
"""
"""
# 엑셀 데이터 조회하여 단어 수 세기
from openpyxl import load_workbook

read_workbook = load_workbook("./blue.xlsx")
read_cell = read_workbook.active
write_list = []
result_crawl = []

for i in range(1, 150) :
    write_list.append( read_cell.cell(i,1).value )  

for i in write_list:
    #print( kkma.nouns( i ))
    #result_crawl += kkma.nouns( i )
    result_crawl += [ s for s in kkma.nouns( i ) if len(s) > 1 ]
    
#print(result_crawl )
result_data = collections.Counter(result_crawl).most_common(20) 
#print( result_data )

# 리스트에서 특정 문자가 포함된 제목 가져오기
#matching = [s for s in write_list if "텔" in s] 
#print( matching )
"""

"""
# 데이터시각화
import numpy as np
import matplotlib.pyplot as plt

plt.bar( ["Andy","Chris","Chalie","Bobby"], [70,42,120,80])
plt.title( "Team Weight" )
plt.xlabel( "Name" )
plt.ylabel( "Weight" )

plt.show()
"""

"""
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import font_manager, rc
font_name = font_manager.FontProperties(fname="c:/Windows/Fonts/malgun.ttf").get_name()
rc('font', family=font_name)
   
   
result_string = []
result_cnt = []
for s, i in result_data :
    result_string.append(s)
    result_cnt.append(i)

index = np.arange( len(result_string))
print( index )
plt.bar( index, result_cnt ) 
plt.xticks( index, result_string )   
plt.xticks( rotation = 45 )
plt.show()
"""

# wordcloud 사용
# conda install -c conda-forge wordcloud

import matplotlib.pyplot as plt
from wordcloud import WordCloud
"""
text = open('2003(No).txt').read()
wc = WordCloud(font_path='./malgun.ttf', background_color="white").generate(text)
print(wc.words_)

#plt.figure( figsize=(12,12) )
#plt.figure()
#plt.imshow( wordcloud, interpolation = "bilinear" )
#plt.imshow( wordcloud )

# Display the generated image:
# the matplotlib way:

fig = plt.figure()
plt.imshow(wc, interpolation='bilinear')
plt.axis('off')
plt.show() 
#plt.savefig('./wordcloud_ex1.svg')
"""
"""
# 청원게시판 내용 기반 워드클라우드
#print( result_crawl )
final_text = ""
for i, s in enumerate(result_crawl) :
    #print( result_crawl[inx] )
    final_text = final_text + " " + s

#print( final_text )

wc = WordCloud(font_path='./malgun.ttf', background_color="white").generate(final_text)

fig = plt.figure( figsize=(14,14))
plt.imshow(wc, interpolation='bilinear')
plt.axis('off')
plt.show() 
"""

# 취임사 내용 기반 워드클라우드

# 1) 워드 클라우드 기능 이용
file_name = ['2003(No).txt', '2008(Lee).txt', '2013(Park).txt', '2017(Mun).txt']

for i in range(0, len( file_name )) :
    print( "파일명 : ", file_name[i] )
    text = open( file_name[i] ).read()
    wc = WordCloud(font_path='./malgun.ttf', background_color="white").generate(text)
    fig = plt.figure( figsize=(14,14))
    plt.imshow(wc, interpolation='bilinear')
    plt.axis('off')
    plt.show()

"""
# 2) konlpy 기능 이용
file_name = ['2003(No).txt', '2008(Lee).txt', '2013(Park).txt', '2017(Mun).txt']
parse_list = []
final_text = ""

for i in range(0, len( file_name )) :
    print( "파일명 : ", file_name[i] )
    text = open( file_name[i] ).read()
    list_sentence = kkma.sentences(text)
    parse_list = []
    final_text = ""
    for i in list_sentence:
        parse_list += [ s for s in kkma.nouns( i ) if len(s) > 1 ]
        for i, s in enumerate(parse_list) :
            final_text = final_text + " " + s
    wc1 = WordCloud(font_path='./malgun.ttf', background_color="white").generate(final_text)
    fig = plt.figure( figsize=(14,14))
    plt.imshow(wc1, interpolation='bilinear')
    plt.axis('off')
    plt.show()
"""    
    
    
    
    
    
    
    
    
    
    
    
    
    